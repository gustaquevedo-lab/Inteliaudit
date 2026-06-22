"""
Endpoints de gestion de hallazgos — ciclo de vida completo.
"""
import json
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel, Field
from sqlalchemy import select, update, or_
from sqlalchemy.ext.asyncio import AsyncSession

from analisis.riesgo import calcular_contingencia, clasificar_riesgo
from db.base import get_db
from db import db as crud
from db.models import Hallazgo, AuditTrail, Usuario
from api.routers.auth import get_current_user, get_current_admin

router = APIRouter(prefix="/auditorias/{auditoria_id}/hallazgos", tags=["hallazgos"])

TRANSICIONES = {
    "pendiente": ["revisado"],
    "revisado": ["aceptado", "descartado"],
    "aceptado": ["regularizado"],
    "descartado": ["revisado"],
    "regularizado": [],
}

class EstadoUpdate(BaseModel):
    estado: str
    notas_auditor: Optional[str] = None

class HallazgoEdit(BaseModel):
    descripcion: Optional[str] = None
    nivel_riesgo: Optional[str] = None
    base_ajuste: Optional[int] = Field(None, ge=0)
    impuesto_omitido: Optional[int] = Field(None, ge=0)
    notas_auditor: Optional[str] = None
    fecha_omision: Optional[str] = None

class EvidenciaInput(BaseModel):
    tipo: str = "rg90"
    referencia_id: str = ""
    descripcion: Optional[str] = None


@router.get("")
async def listar_hallazgos(
    auditoria_id: str,
    impuesto: Optional[str] = None,
    estado: Optional[str] = None,
    nivel_riesgo: Optional[str] = None,
    periodo: Optional[str] = None,
    busqueda: Optional[str] = None,
    ordenar: str = "contingencia_desc",
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    auditoria = await crud.get_auditoria(db, user.firma_id, auditoria_id)
    if not auditoria:
        raise HTTPException(404, "Auditoria no encontrada")

    q = select(Hallazgo).where(
        Hallazgo.auditoria_id == auditoria_id,
        Hallazgo.firma_id == user.firma_id,
    )
    if impuesto:
        q = q.where(Hallazgo.impuesto == impuesto.upper())
    if estado:
        q = q.where(Hallazgo.estado == estado)
    if nivel_riesgo:
        q = q.where(Hallazgo.nivel_riesgo == nivel_riesgo)
    if periodo:
        q = q.where(Hallazgo.periodo == periodo)
    if busqueda:
        q = q.where(or_(
            Hallazgo.descripcion.ilike(f"%{busqueda}%"),
            Hallazgo.tipo_hallazgo.ilike(f"%{busqueda}%"),
        ))

    if ordenar == "contingencia_desc":
        q = q.order_by(Hallazgo.total_contingencia.desc())
    elif ordenar == "contingencia_asc":
        q = q.order_by(Hallazgo.total_contingencia)
    elif ordenar == "periodo":
        q = q.order_by(Hallazgo.periodo.desc())
    else:
        q = q.order_by(Hallazgo.creado_en.desc())

    result = await db.execute(q)
    return [_serializar(h) for h in result.scalars().all()]


@router.get("/export-excel")
async def exportar_hallazgos_excel(
    auditoria_id: str,
    impuesto: Optional[str] = None,
    estado: Optional[str] = None,
    nivel_riesgo: Optional[str] = None,
    periodo: Optional[str] = None,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Genera XLSX con hallazgos filtrados."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    import io

    auditoria = await crud.get_auditoria(db, user.firma_id, auditoria_id)
    if not auditoria:
        raise HTTPException(404, "Auditoria no encontrada")

    q = select(Hallazgo).where(Hallazgo.auditoria_id == auditoria_id, Hallazgo.firma_id == user.firma_id)
    if impuesto: q = q.where(Hallazgo.impuesto == impuesto.upper())
    if estado: q = q.where(Hallazgo.estado == estado)
    if nivel_riesgo: q = q.where(Hallazgo.nivel_riesgo == nivel_riesgo)
    if periodo: q = q.where(Hallazgo.periodo == periodo)
    q = q.order_by(Hallazgo.total_contingencia.desc())

    result = await db.execute(q)
    hallazgos = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Hallazgos"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2E84F0")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    cols = ["#", "Impuesto", "Periodo", "Tipo", "Descripcion", "Art. Legal",
            "Base Ajuste", "Impuesto Omitido", "Multa", "Intereses", "Total", "Riesgo", "Estado"]
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for idx, h in enumerate(hallazgos, 1):
        row = [
            idx, h.impuesto, h.periodo, h.tipo_hallazgo, h.descripcion[:100],
            h.articulo_legal[:80], h.base_ajuste, h.impuesto_omitido,
            h.multa_estimada, h.intereses_estimados, h.total_contingencia,
            h.nivel_riesgo, h.estado,
        ]
        for i, v in enumerate(row, 1):
            cell = ws.cell(row=idx + 1, column=i, value=v)
            cell.border = border

    ws.column_dimensions["A"].width = 5
    for c in "BC": ws.column_dimensions[c].width = 14
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 40
    for c in "GHIJK": ws.column_dimensions[c].width = 16
    for c in "LM": ws.column_dimensions[c].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=hallazgos_{auditoria_id[:8]}.xlsx"})


@router.post("/estado-batch")
async def cambiar_estado_batch(
    auditoria_id: str,
    body: dict,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Cambia estado de multiples hallazgos a la vez."""
    ids = body.get("ids", [])
    nuevo_estado = body.get("estado", "")
    notas = body.get("notas_auditor")

    if user.rol not in ("super_admin", "admin", "auditor_senior"):
        raise HTTPException(403, "Solo admin o auditor senior puede cambiar estados")

    if nuevo_estado not in ("revisado", "aceptado", "descartado"):
        raise HTTPException(400, f"Estado invalido: {nuevo_estado}")

    count = 0
    for hid in ids:
        result = await db.execute(
            select(Hallazgo).where(Hallazgo.id == hid, Hallazgo.firma_id == user.firma_id)
        )
        h = result.scalar_one_or_none()
        if not h:
            continue
        trans = TRANSICIONES.get(h.estado, [])
        if nuevo_estado not in trans:
            continue
        vals = {"estado": nuevo_estado}
        if notas:
            vals["notas_auditor"] = notas
        await db.execute(update(Hallazgo).where(Hallazgo.id == hid).values(**vals))
        count += 1

    if count > 0:
        await crud.log_trail(db, firma_id=user.firma_id, usuario_id=user.id,
            accion=f"Estado batch: {count} hallazgos → {nuevo_estado}",
            modulo="hallazgos", auditoria_id=auditoria_id, datos={"ids": ids, "estado": nuevo_estado})
        await db.commit()

    return {"ok": True, "actualizados": count}


@router.patch("/{hallazgo_id}/estado")
async def cambiar_estado_hallazgo(
    auditoria_id: str,
    hallazgo_id: str,
    body: EstadoUpdate,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Cambia estado de un hallazgo con validacion de transiciones."""
    if user.rol not in ("super_admin", "admin", "auditor_senior"):
        raise HTTPException(403, "Solo admin o auditor senior puede cambiar estados")

    result = await db.execute(
        select(Hallazgo).where(Hallazgo.id == hallazgo_id, Hallazgo.firma_id == user.firma_id)
    )
    h = result.scalar_one_or_none()
    if not h:
        raise HTTPException(404, "Hallazgo no encontrado")

    trans = TRANSICIONES.get(h.estado, [])
    if body.estado not in trans:
        raise HTTPException(400, f"Transicion invalida: {h.estado} → {body.estado}. Permitidas: {trans}")

    vals = {"estado": body.estado}
    if body.notas_auditor is not None:
        vals["notas_auditor"] = body.notas_auditor

    await db.execute(update(Hallazgo).where(Hallazgo.id == hallazgo_id).values(**vals))

    await crud.log_trail(db, firma_id=user.firma_id, usuario_id=user.id,
        accion=f"Hallazgo {body.estado}: {h.tipo_hallazgo}",
        modulo="hallazgos", auditoria_id=auditoria_id,
        datos={"hallazgo_id": hallazgo_id, "estado_anterior": h.estado, "estado_nuevo": body.estado})

    if body.estado == "aceptado":
        # Verificar si quedan pendientes
        restantes = await db.execute(
            select(Hallazgo).where(
                Hallazgo.auditoria_id == auditoria_id,
                Hallazgo.firma_id == user.firma_id,
                Hallazgo.estado == "pendiente",
            )
        )
        if restantes.scalars().all():
            await db.commit()
            return {"ok": True, "advertencia": "Aun hay hallazgos pendientes de revision"}

    await db.commit()
    return {"ok": True, "estado": body.estado}


@router.patch("/{hallazgo_id}")
async def editar_hallazgo(
    auditoria_id: str,
    hallazgo_id: str,
    body: HallazgoEdit,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Edita campos de un hallazgo y recalcula contingencia automaticamente."""
    if user.rol not in ("super_admin", "admin", "auditor_senior"):
        raise HTTPException(403, "Solo admin o auditor senior puede editar")

    result = await db.execute(
        select(Hallazgo).where(Hallazgo.id == hallazgo_id, Hallazgo.firma_id == user.firma_id)
    )
    h = result.scalar_one_or_none()
    if not h:
        raise HTTPException(404, "Hallazgo no encontrado")

    vals = {k: v for k, v in body.model_dump(exclude_none=True).items()}

    if "impuesto_omitido" in vals or "base_ajuste" in vals:
        imp = vals.get("impuesto_omitido", h.impuesto_omitido)
        fecha_om = vals.get("fecha_omision") or h.fecha_emision or h.periodo + "-20"
        cont = calcular_contingencia(imp, fecha_om)
        vals["multa_estimada"] = cont["multa_estimada"]
        vals["intereses_estimados"] = cont["intereses_estimados"]
        vals["total_contingencia"] = cont["total_contingencia"]
        vals["nivel_riesgo"] = clasificar_riesgo(cont["total_contingencia"])

    if vals:
        await db.execute(update(Hallazgo).where(Hallazgo.id == hallazgo_id).values(**vals))
        await crud.log_trail(db, firma_id=user.firma_id, usuario_id=user.id,
            accion=f"Hallazgo editado: {hallazgo_id[:8]}",
            modulo="hallazgos", auditoria_id=auditoria_id, datos={"campos": list(vals.keys())})
        await db.commit()

    result2 = await db.execute(select(Hallazgo).where(Hallazgo.id == hallazgo_id))
    return _serializar(result2.scalar_one())


@router.post("/{hallazgo_id}/evidencias")
async def agregar_evidencia(
    auditoria_id: str,
    hallazgo_id: str,
    body: EvidenciaInput,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Vincula un comprobante como evidencia del hallazgo."""
    result = await db.execute(
        select(Hallazgo).where(Hallazgo.id == hallazgo_id, Hallazgo.firma_id == user.firma_id)
    )
    h = result.scalar_one_or_none()
    if not h:
        raise HTTPException(404, "Hallazgo no encontrado")

    evidencias = json.loads(h.evidencias or "[]")
    nueva = {"tipo": body.tipo, "ref": body.referencia_id, "desc": body.descripcion or ""}
    evidencias.append(nueva)

    await db.execute(
        update(Hallazgo).where(Hallazgo.id == hallazgo_id).values(evidencias=json.dumps(evidencias))
    )
    await db.commit()
    return {"ok": True, "evidencias": evidencias}


@router.get("/{hallazgo_id}/trail")
async def historial_hallazgo(
    auditoria_id: str,
    hallazgo_id: str,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Historial de cambios del hallazgo desde audit_trail."""
    result = await db.execute(
        select(AuditTrail)
        .where(
            AuditTrail.firma_id == user.firma_id,
            AuditTrail.auditoria_id == auditoria_id,
            AuditTrail.modulo == "hallazgos",
        )
        .order_by(AuditTrail.timestamp.desc())
        .limit(20)
    )
    rows = result.scalars().all()
    items = []
    for r in rows:
        try:
            datos = json.loads(r.datos_json or "{}")
            if datos.get("hallazgo_id") == hallazgo_id or hallazgo_id in str(datos.get("ids", [])):
                items.append({
                    "accion": r.accion,
                    "timestamp": r.timestamp.isoformat() if r.timestamp else None,
                    "detalle": r.detalle,
                    "datos": datos,
                })
        except (json.JSONDecodeError, TypeError):
            pass
    return items


@router.post("", status_code=201)
async def crear_hallazgo_manual(
    auditoria_id: str,
    body: dict,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    auditoria = await crud.get_auditoria(db, user.firma_id, auditoria_id)
    if not auditoria:
        raise HTTPException(404, "Auditoria no encontrada")

    impuesto_omitido = int(body.get("impuesto_omitido", 0))
    fecha_om = body.get("fecha_omision", body.get("periodo", "") + "-20")
    cont = calcular_contingencia(impuesto_omitido, fecha_om) if impuesto_omitido > 0 else {"multa_estimada": 0, "intereses_estimados": 0, "total_contingencia": 0}

    h = await crud.crear_hallazgo(
        db, firma_id=user.firma_id, auditoria_id=auditoria_id,
        impuesto=body.get("impuesto", "IVA"),
        periodo=body.get("periodo", ""),
        tipo_hallazgo=body.get("tipo_hallazgo", ""),
        descripcion=body.get("descripcion", ""),
        articulo_legal=body.get("articulo_legal", ""),
        base_ajuste=int(body.get("base_ajuste", 0)),
        impuesto_omitido=impuesto_omitido,
        multa_estimada=cont["multa_estimada"],
        intereses_estimados=cont["intereses_estimados"],
        nivel_riesgo=clasificar_riesgo(cont["total_contingencia"], auditoria.materialidad) if impuesto_omitido > 0 else "bajo",
        evidencias=body.get("evidencias", []),
        creado_por="auditor",
    )
    await db.commit()
    return _serializar(h)


# ============================================================
#  Narrativa IA (Gemini / Claude)
# ============================================================

@router.post("/{hallazgo_id}/generar-narrativa")
async def generar_narrativa_hallazgo_endpoint(
    auditoria_id: str,
    hallazgo_id: str,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from config.plans import PLAN_ALIAS_MAP, get_plan
    from db.models import Firma
    firma_result = await db.execute(select(Firma).where(Firma.id == user.firma_id))
    firma = firma_result.scalar_one_or_none()
    if not firma:
        raise HTTPException(404, "Firma no encontrada")
    plan_key = PLAN_ALIAS_MAP.get(firma.plan, firma.plan)
    plan_cfg = get_plan(plan_key)
    if not plan_cfg.tiene_ia:
        raise HTTPException(403, "Narrativa IA disponible en plan Pro y Enterprise")

    from analisis.rate_limiter import check_ia_rate_limit, increment_ia_usage
    puede, msg, usos = check_ia_rate_limit(user.firma_id, firma.plan)
    if not puede:
        raise HTTPException(429, msg)

    result = await db.execute(select(Hallazgo).where(Hallazgo.id == hallazgo_id, Hallazgo.firma_id == user.firma_id))
    hallazgo = result.scalar_one_or_none()
    if not hallazgo:
        raise HTTPException(404, "Hallazgo no encontrado")
    auditoria = await crud.get_auditoria(db, user.firma_id, auditoria_id)
    cliente = await crud.get_cliente(db, user.firma_id, id=auditoria.cliente_id) if auditoria else None
    h_dict = _serializar(hallazgo)
    from analisis.claude_analisis import generar_narrativa_hallazgo as gn
    narrativa = gn(hallazgo=h_dict, cliente={"razon_social": cliente.razon_social, "ruc": cliente.ruc, "actividad_principal": cliente.actividad_principal or ""}, auditoria={"periodo_desde": auditoria.periodo_desde, "periodo_hasta": auditoria.periodo_hasta})
    await db.execute(update(Hallazgo).where(Hallazgo.id == hallazgo_id).values(descripcion=narrativa, sugerencia_ai=True))
    await db.commit()
    from analisis.rate_limiter import increment_ia_usage as inc
    inc(user.firma_id)
    return {"ok": True, "narrativa": narrativa}


@router.post("/generar-narrativas-batch")
async def generar_narrativas_batch(
    auditoria_id: str,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from config.plans import PLAN_ALIAS_MAP, get_plan
    from db.models import Firma
    firma_result = await db.execute(select(Firma).where(Firma.id == user.firma_id))
    firma = firma_result.scalar_one_or_none()
    if not firma:
        raise HTTPException(404, "Firma no encontrada")
    plan_key = PLAN_ALIAS_MAP.get(firma.plan, firma.plan)
    plan_cfg = get_plan(plan_key)
    if not plan_cfg.tiene_ia:
        raise HTTPException(403, "Narrativa IA disponible en plan Pro y Enterprise")
    result = await db.execute(select(Hallazgo).where(Hallazgo.auditoria_id == auditoria_id, Hallazgo.firma_id == user.firma_id, Hallazgo.estado != "descartado").order_by(Hallazgo.total_contingencia.desc()))
    hallazgos = result.scalars().all()
    if not hallazgos:
        raise HTTPException(400, "No hay hallazgos para narrar")
    auditoria = await crud.get_auditoria(db, user.firma_id, auditoria_id)
    cliente = await crud.get_cliente(db, user.firma_id, id=auditoria.cliente_id) if auditoria else None
    from analisis.claude_analisis import generar_narrativa_hallazgo as gn
    from analisis.rate_limiter import check_ia_rate_limit, increment_ia_usage as inc
    puede, msg, _ = check_ia_rate_limit(user.firma_id, firma.plan)
    if not puede:
        raise HTTPException(429, msg)
    narrados = 0
    errores = 0
    for h in hallazgos:
        try:
            h_dict = _serializar(h)
            narrativa = gn(hallazgo=h_dict, cliente={"razon_social": cliente.razon_social, "ruc": cliente.ruc, "actividad_principal": cliente.actividad_principal or ""}, auditoria={"periodo_desde": auditoria.periodo_desde, "periodo_hasta": auditoria.periodo_hasta})
            await db.execute(update(Hallazgo).where(Hallazgo.id == h.id).values(descripcion=narrativa, sugerencia_ai=True))
            inc(user.firma_id)
            narrados += 1
        except Exception:
            errores += 1
    await db.commit()
    return {"ok": True, "narrados": narrados, "errores": errores, "total": len(hallazgos)}


def _serializar(h: Hallazgo) -> dict:
    return {
        "id": h.id,
        "impuesto": h.impuesto,
        "periodo": h.periodo,
        "tipo_hallazgo": h.tipo_hallazgo,
        "descripcion": h.descripcion,
        "descripcion_tecnica": h.descripcion_tecnica,
        "articulo_legal": h.articulo_legal,
        "base_ajuste": h.base_ajuste,
        "impuesto_omitido": h.impuesto_omitido,
        "multa_estimada": h.multa_estimada,
        "intereses_estimados": h.intereses_estimados,
        "total_contingencia": h.total_contingencia,
        "nivel_riesgo": h.nivel_riesgo,
        "estado": h.estado,
        "evidencias": json.loads(h.evidencias or "[]"),
        "notas_auditor": h.notas_auditor,
        "creado_por": h.creado_por,
        "creado_en": h.creado_en.isoformat() if h.creado_en else None,
    }


global_router = APIRouter(prefix="/hallazgos", tags=["hallazgos"])

@global_router.get("")
async def listar_todos_hallazgos(
    cliente_id: Optional[str] = None,
    impuesto: Optional[str] = None,
    estado: Optional[str] = None,
    nivel_riesgo: Optional[str] = None,
    periodo: Optional[str] = None,
    busqueda: Optional[str] = None,
    user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from db.models import Auditoria, Cliente
    q = select(Hallazgo, Cliente.razon_social, Cliente.ruc).join(
        Auditoria, Hallazgo.auditoria_id == Auditoria.id
    ).join(
        Cliente, Auditoria.cliente_id == Cliente.id
    ).where(Hallazgo.firma_id == user.firma_id)

    if cliente_id:
        q = q.where(Auditoria.cliente_id == cliente_id)
    if impuesto:
        q = q.where(Hallazgo.impuesto == impuesto.upper())
    if estado:
        q = q.where(Hallazgo.estado == estado)
    if nivel_riesgo:
        q = q.where(Hallazgo.nivel_riesgo == nivel_riesgo)
    if periodo:
        q = q.where(Hallazgo.periodo == periodo)
    if busqueda:
        q = q.where(or_(
            Hallazgo.descripcion.ilike(f"%{busqueda}%"),
            Hallazgo.tipo_hallazgo.ilike(f"%{busqueda}%"),
            Cliente.razon_social.ilike(f"%{busqueda}%"),
        ))

    q = q.order_by(Hallazgo.total_contingencia.desc())
    result = await db.execute(q)
    rows = result.all()
    
    return [
        {
            **_serializar(h),
            "cliente_nombre": razon_social,
            "cliente_ruc": ruc,
            "auditoria_id": h.auditoria_id,
        }
        for h, razon_social, ruc in rows
    ]

