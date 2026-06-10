"""
Genera archivos de fixtures para los tests.
Ejecutar: python tests/fixtures/generate_fixtures.py
"""
from pathlib import Path
import openpyxl

FIXTURES_DIR = Path(__file__).parent


def generar_rg90_compras():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compras"

    headers = [
        "RUC Proveedor", "Razón Social", "Timbrado", "Establecimiento",
        "Punto Expedición", "Nro. Comprobante", "CDC", "Tipo Comprobante",
        "Fecha Emisión", "Gravado 10%", "Gravado 5%", "Exento",
        "IVA 10%", "IVA 5%", "Total IVA", "Total Comprobante",
    ]
    ws.append(headers)

    ws.append([
        "80055555-1", "Proveedor Activo SA", "12345678", "001", "001",
        "0000001", "12345678901234567890123456789012345678901234", "1",
        "15/03/2024", 10000000, 0, 0, 1000000, 0, 1000000, 11000000,
    ])
    ws.append([
        "80066666-2", "Proveedor Inactivo SA", "12345679", "001", "001",
        "0000002", "99999999999999999999999999999999999999999999", "1",
        "20/03/2024", 5000000, 0, 0, 500000, 0, 500000, 5500000,
    ])
    ws.append([
        "80077777-3", "Sin CDC SA", "12345680", "001", "001",
        "0000003", "", "1",
        "25/03/2024", 3000000, 0, 0, 300000, 0, 300000, 3300000,
    ])

    path = FIXTURES_DIR / "rg90_compras.xlsx"
    wb.save(path)
    return path


def generar_rg90_ventas():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"

    headers = [
        "RUC Cliente", "Razón Social", "Timbrado", "Establecimiento",
        "Punto Expedición", "Nro. Comprobante", "CDC", "Tipo Comprobante",
        "Fecha Emisión", "Gravado 10%", "Gravado 5%", "Exento",
        "IVA 10%", "IVA 5%", "Total IVA", "Total Comprobante",
    ]
    ws.append(headers)

    ws.append([
        "80088888-4", "Cliente A SA", "87654321", "001", "001",
        "0000100", "44444444444444444444444444444444444444444444", "1",
        "10/03/2024", 20000000, 0, 0, 2000000, 0, 2000000, 22000000,
    ])
    ws.append([
        "80099999-5", "Cliente B SA", "87654322", "001", "001",
        "0000101", "55555555555555555555555555555555555555555555", "1",
        "18/03/2024", 15000000, 0, 0, 1500000, 0, 1500000, 16500000,
    ])

    path = FIXTURES_DIR / "rg90_ventas.xlsx"
    wb.save(path)
    return path


def generar_hechauka():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Informacion Recibida"

    headers = [
        "RUC Informante", "Nombre Informante", "Tipo Operacion",
        "Nro. Comprobante", "Fecha Comprobante", "Monto Operacion",
        "IVA Operacion", "Retencion IVA", "Retencion IRE",
    ]
    ws.append(headers)

    ws.append([
        "80088888-4", "Cliente A SA", "compra",
        "001-001-0000100", "10/03/2024", 22000000, 2000000, 600000, 0,
    ])
    ws.append([
        "80099999-5", "Cliente B SA", "compra",
        "001-001-0000101", "18/03/2024", 16500000, 1500000, 450000, 0,
    ])
    ws.append([
        "80011111-7", "Cliente Omitido SA", "compra",
        "001-001-0000999", "28/03/2024", 8000000, 800000, 240000, 0,
    ])

    path = FIXTURES_DIR / "hechauka.xlsx"
    wb.save(path)
    return path


def generar_sifen_xml():
    xml = """<?xml version="1.0" encoding="UTF-8"?>
<rDE xmlns="http://ekuatia.set.gov.py/sifen/xsd">
  <DE>
    <gTimb>
      <dTiDE>1</dTiDE>
      <dNumTim>12345678</dNumTim>
      <dEst>001</dEst>
      <dPunExp>001</dPunExp>
      <dNumDoc>0000001</dNumDoc>
    </gTimb>
    <gDatGralOpe>
      <dFeEmiDE>2024-03-15T10:30:00</dFeEmiDE>
      <gDatRec>
        <dRucRec>80012345-6</dRucRec>
        <dNomRec>Comercial Guarani SA</dNomRec>
      </gDatRec>
    </gDatGralOpe>
    <gEmis>
      <dRucEm>80055555-1</dRucEm>
      <dNomEmi>Proveedor Activo SA</dNomEmi>
    </gEmis>
    <gTotSub>
      <dTotGravOp10>10000000</dTotGravOp10>
      <dTotGravOp5>0</dTotGravOp5>
      <dTotExe>0</dTotExe>
      <dTotIVA>1000000</dTotIVA>
      <dTotGe>11000000</dTotGe>
    </gTotSub>
    <Id>12345678901234567890123456789012345678901234</Id>
  </DE>
</rDE>"""
    path = FIXTURES_DIR / "sifen_de.xml"
    path.write_text(xml, encoding="utf-8")
    return path


if __name__ == "__main__":
    print(f"Generando fixtures en {FIXTURES_DIR}...")
    print(f"  RG90 Compras: {generar_rg90_compras()}")
    print(f"  RG90 Ventas:  {generar_rg90_ventas()}")
    print(f"  HECHAUKA:     {generar_hechauka()}")
    print(f"  SIFEN XML:    {generar_sifen_xml()}")
    print("Hecho.")
