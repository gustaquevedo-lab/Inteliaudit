"""
Parser de archivos XLSX RG90 descargados desde Marangatú.
El formato está definido por RG 69/2020 y RG 90/2021.
"""
from pathlib import Path
from typing import Optional

import openpyxl
from rich.console import Console

console = Console()

# Columnas del XLSX RG90 según formato SET
# Los nombres exactos pueden variar según versión — ajustar si cambia
COLUMNAS_COMPRAS = {
    "RUC Proveedor": "ruc_contraparte",
    "Razón Social": "nombre_contraparte",
    "Timbrado": "timbrado",
    "Establecimiento": "establecimiento",
    "Punto Expedición": "punto_expedicion",
    "Nro. Comprobante": "nro_comprobante",
    "CDC": "cdc",
    "Tipo Comprobante": "tipo_comprobante",
    "Fecha Emisión": "fecha_emision",
    "Gravado 10%": "base_gravada_10",
    "Gravado 5%": "base_gravada_5",
    "Exento": "monto_exento",
    "IVA 10%": "iva_10",
    "IVA 5%": "iva_5",
    "Total IVA": "iva_total",
    "Total Comprobante": "total_comprobante",
}

COLUMNAS_VENTAS = {
    "RUC Cliente": "ruc_contraparte",
    "Razón Social": "nombre_contraparte",
    "Timbrado": "timbrado",
    "Establecimiento": "establecimiento",
    "Punto Expedición": "punto_expedicion",
    "Nro. Comprobante": "nro_comprobante",
    "CDC": "cdc",
    "Tipo Comprobante": "tipo_comprobante",
    "Fecha Emisión": "fecha_emision",
    "Gravado 10%": "base_gravada_10",
    "Gravado 5%": "base_gravada_5",
    "Exento": "monto_exento",
    "IVA 10%": "iva_10",
    "IVA 5%": "iva_5",
    "Total IVA": "iva_total",
    "Total Comprobante": "total_comprobante",
}


def parsear_rg90(
    archivo: Path,
    cliente_id: str,
    periodo: str,
    auditoria_id: Optional[str] = None,
) -> list[dict]:
    """
    Parsea un XLSX de RG90 y retorna lista de dicts listos para insertar en DB.

    Args:
        archivo: Path al archivo XLSX
        cliente_id: ID (UUID) del contribuyente auditado
        periodo: Período YYYY-MM al que corresponde el archivo
        auditoria_id: ID de la auditoría activa (opcional)

    Returns:
        Lista de dicts con estructura de la tabla rg90
    """
    console.print(f"[blue]Parser RG90:[/] procesando {archivo.name}...")
    wb = openpyxl.load_workbook(archivo, data_only=True)

    registros: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        tipo = _detectar_tipo(sheet_name)
        if tipo is None:
            continue

        mapa = COLUMNAS_COMPRAS if tipo == "compra" else COLUMNAS_VENTAS
        headers = _leer_headers(ws)
        col_map = _mapear_columnas(headers, mapa)

        if not col_map:
            console.print(f"[yellow]⚠[/] Hoja '{sheet_name}': columnas no reconocidas, omitiendo.")
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            if _fila_vacia(row):
                continue
            registro = _parsear_fila(row, col_map, tipo, cliente_id, periodo, auditoria_id, archivo.name)
            if registro:
                registros.append(registro)

    console.print(f"[green]✓[/] RG90 {archivo.name}: {len(registros)} comprobantes.")
    return registros


def _detectar_tipo(sheet_name: str) -> Optional[str]:
    """Detecta si la hoja es de compras o ventas por el nombre."""
    nombre = sheet_name.lower()
    if "compra" in nombre:
        return "compra"
    if "venta" in nombre:
        return "venta"
    return None


def _leer_headers(ws) -> list[str]:
    """Lee la primera fila como headers."""
    return [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]


def _mapear_columnas(headers: list[str], mapa: dict) -> dict[str, int]:
    """
    Crea un dict {campo_db: indice_columna} basado en los headers del XLSX.
    Tolerante a mayúsculas/minúsculas y espacios extra.
    """
    headers_norm = {h.lower().strip(): i for i, h in enumerate(headers)}
    col_map = {}
    for col_xlsx, campo_db in mapa.items():
        idx = headers_norm.get(col_xlsx.lower().strip())
        if idx is not None:
            col_map[campo_db] = idx
    return col_map


def _parsear_fila(
    row: tuple,
    col_map: dict[str, int],
    tipo: str,
    cliente_id: str,
    periodo: str,
    auditoria_id: Optional[str],
    fuente_archivo: str,
) -> Optional[dict]:
    """Convierte una fila del XLSX en un dict para la tabla rg90."""

    def get(campo: str):
        idx = col_map.get(campo)
        return row[idx] if idx is not None and idx < len(row) else None

    ruc_contraparte = _limpiar_ruc(get("ruc_contraparte"))
    nro_comprobante = str(get("nro_comprobante") or "").strip()
    fecha_emision = _normalizar_fecha(get("fecha_emision"))

    if not ruc_contraparte or not nro_comprobante or not fecha_emision:
        return None

    return {
        "auditoria_id": auditoria_id,
        "cliente_id": cliente_id,
        "periodo": periodo,
        "tipo": tipo,
        "ruc_contraparte": ruc_contraparte,
        "nombre_contraparte": str(get("nombre_contraparte") or "").strip() or None,
        "timbrado": str(get("timbrado") or "").strip() or None,
        "establecimiento": str(get("establecimiento") or "").strip() or None,
        "punto_expedicion": str(get("punto_expedicion") or "").strip() or None,
        "nro_comprobante": nro_comprobante,
        "cdc": _limpiar_cdc(get("cdc")),
        "tipo_comprobante": str(get("tipo_comprobante") or "").strip() or None,
        "fecha_emision": fecha_emision,
        "base_gravada_10": _int_pyg(get("base_gravada_10")),
        "base_gravada_5": _int_pyg(get("base_gravada_5")),
        "monto_exento": _int_pyg(get("monto_exento")),
        "iva_10": _int_pyg(get("iva_10")),
        "iva_5": _int_pyg(get("iva_5")),
        "iva_total": _int_pyg(get("iva_total")),
        "total_comprobante": _int_pyg(get("total_comprobante")),
        "fuente_archivo": fuente_archivo,
    }


# --------------------------------------------------------
#  Helpers de normalización
# --------------------------------------------------------

def _fila_vacia(row: tuple) -> bool:
    return all(v is None or str(v).strip() == "" for v in row)


def _limpiar_ruc(valor) -> Optional[str]:
    """Normaliza RUC: elimina espacios, garantiza formato XXXXXXXX-D."""
    if not valor:
        return None
    ruc = str(valor).strip().replace(" ", "")
    return ruc if ruc else None


def _limpiar_cdc(valor) -> Optional[str]:
    """Valida y retorna CDC de 44 dígitos, o None si no es válido."""
    if not valor:
        return None
    cdc = str(valor).strip().replace(" ", "")
    if len(cdc) == 44 and cdc.isdigit():
        return cdc
    return None


def _normalizar_fecha(valor) -> Optional[str]:
    """Convierte varios formatos de fecha a YYYY-MM-DD."""
    if not valor:
        return None
    from datetime import date, datetime

    if isinstance(valor, (date, datetime)):
        return valor.strftime("%Y-%m-%d")
    s = str(valor).strip()
    # Intentar formatos comunes del XLSX paraguayo
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            from datetime import datetime as dt
            return dt.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s  # retornar tal cual si no se pudo parsear


def _int_pyg(valor) -> int:
    """Convierte valor a entero PYG (sin decimales). None → 0."""
    if valor is None:
        return 0
    try:
        return int(round(float(str(valor).replace(",", "").replace(".", "") or "0")))
    except (ValueError, TypeError):
        return 0
